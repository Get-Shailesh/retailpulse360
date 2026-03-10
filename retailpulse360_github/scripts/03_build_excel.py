"""
RetailPulse 360 — Professional Excel Management Report Builder
Creates a multi-sheet KPI workbook using openpyxl
"""
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, CellIsRule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
import os

DATA = "/home/claude/retailpulse360/data"
OUT  = "/home/claude/retailpulse360/outputs"
os.makedirs(OUT, exist_ok=True)

# ── Load data ──────────────────────────────────────────────────
trans     = pd.read_csv(f"{DATA}/transactions.csv", parse_dates=["date"])
stores    = pd.read_csv(f"{DATA}/stores.csv")
products  = pd.read_csv(f"{DATA}/products.csv")
customers = pd.read_csv(f"{DATA}/customers.csv")
rfm       = pd.read_csv(f"{DATA}/rfm_scores.csv")
churn     = pd.read_csv(f"{DATA}/churn_scores.csv")

print("Building Excel Report...")

wb = Workbook()

# ── Colour palette ─────────────────────────────────────────────
C = {
    "dark_bg":   "0A0A1A",
    "header":    "001F3F",
    "accent":    "00C97A",
    "accent2":   "4A90D9",
    "danger":    "E53E3E",
    "warning":   "F6C90E",
    "light_bg":  "F0F4F8",
    "white":     "FFFFFF",
    "gray":      "6B7280",
    "border":    "D1D5DB",
    "gold":      "F6AD55",
    "green_bg":  "ECFDF5",
    "red_bg":    "FEF2F2",
    "blue_bg":   "EFF6FF",
}

def bold_white(size=11):
    return Font(name="Arial", bold=True, color=C["white"], size=size)
def bold_dark(size=11):
    return Font(name="Arial", bold=True, color=C["dark_bg"], size=size)
def normal(size=10, color="1A1A1A"):
    return Font(name="Arial", size=size, color=color)
def hdr_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)
def thin_border():
    s = Side(style='thin', color=C["border"])
    return Border(left=s, right=s, top=s, bottom=s)
def center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)
def right():
    return Alignment(horizontal='right', vertical='center')

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def write_header_row(ws, row, cols_data, fill_color=None):
    for col, (text, width) in enumerate(cols_data, 1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font  = bold_white(10)
        cell.fill  = hdr_fill(fill_color or C["header"])
        cell.alignment = center()
        cell.border = thin_border()
        set_col_width(ws, col, width)

def write_data_row(ws, row, values, even=True, formats=None):
    bg = C["light_bg"] if even else C["white"]
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font   = normal(9)
        cell.fill   = hdr_fill(bg)
        cell.border = thin_border()
        cell.alignment = right() if isinstance(val, (int, float)) else Alignment(vertical='center')
        if formats and col-1 < len(formats) and formats[col-1]:
            cell.number_format = formats[col-1]

# ══════════════════════════════════════════════════════════════
# SHEET 1: EXECUTIVE SUMMARY
# ══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "📊 Executive Summary"
ws1.sheet_view.showGridLines = False
ws1.row_dimensions[1].height = 50

# Title banner
ws1.merge_cells("A1:J1")
tc = ws1["A1"]
tc.value = "NOVAMART — RETAILPULSE 360  |  EXECUTIVE SUMMARY  |  FY 2024"
tc.font  = Font(name="Arial", bold=True, color=C["white"], size=16)
tc.fill  = hdr_fill(C["dark_bg"])
tc.alignment = center()

ws1.merge_cells("A2:J2")
ws1["A2"].value = "Confidential — For Executive Team Distribution Only"
ws1["A2"].font  = Font(name="Arial", italic=True, color=C["gray"], size=9)
ws1["A2"].fill  = hdr_fill("1A1A2E")
ws1["A2"].alignment = center()

# KPI section header
ws1.row_dimensions[4].height = 28
ws1.merge_cells("A4:J4")
ws1["A4"].value = "KEY PERFORMANCE INDICATORS — FY2024"
ws1["A4"].font  = bold_white(11)
ws1["A4"].fill  = hdr_fill(C["accent2"])
ws1["A4"].alignment = center()

# Compute KPIs
fy24 = trans[trans['date'].dt.year == 2024]
fy23 = trans[trans['date'].dt.year == 2023]
total_rev_24  = fy24['total_revenue'].sum()
total_rev_23  = fy23['total_revenue'].sum()
gp_24         = fy24['gross_profit'].sum()
gp_23         = fy23['gross_profit'].sum()
txns_24       = len(fy24)
uniq_custs_24 = fy24['customer_id'].nunique()
aov_24        = fy24['total_revenue'].mean()
aov_23        = fy23['total_revenue'].mean()
gp_margin_24  = gp_24 / total_rev_24

kpi_labels = [
    ("Total Revenue",      f"${total_rev_24/1e6:.1f}M",  f"{(total_rev_24-total_rev_23)/total_rev_23*100:+.1f}%", C["accent"]),
    ("Gross Profit",       f"${gp_24/1e6:.1f}M",         f"{(gp_24-gp_23)/gp_23*100:+.1f}%",                     C["accent"]),
    ("GP Margin",          f"{gp_margin_24*100:.1f}%",    f"vs {gp_23/total_rev_23*100:.1f}% PY",                C["accent2"]),
    ("Transactions",       f"{txns_24:,}",                f"{(txns_24-len(fy23))/len(fy23)*100:+.1f}% YoY",       C["accent2"]),
    ("Active Customers",   f"{uniq_custs_24:,}",          "Unique FY24",                                          C["gold"]),
    ("Avg Order Value",    f"${aov_24:.2f}",              f"{(aov_24-aov_23)/aov_23*100:+.1f}% YoY",             C["gold"]),
]

# KPI cards in row 6–9 (2 rows per card, 3 cards per row × 2 rows)
for i, (label, value, change, color) in enumerate(kpi_labels):
    col_start = (i % 3) * 3 + 1  # cols 1,4,7
    base_row  = 6 if i < 3 else 11

    # Label row
    ws1.merge_cells(start_row=base_row,   start_column=col_start, end_row=base_row,   end_column=col_start+2)
    lc = ws1.cell(row=base_row, column=col_start, value=label)
    lc.font = Font(name="Arial", bold=True, color=C["white"], size=9)
    lc.fill = hdr_fill(color)
    lc.alignment = center()
    ws1.row_dimensions[base_row].height = 18

    # Value row
    ws1.merge_cells(start_row=base_row+1, start_column=col_start, end_row=base_row+2, end_column=col_start+2)
    vc = ws1.cell(row=base_row+1, column=col_start, value=value)
    vc.font = Font(name="Arial", bold=True, color=color, size=18)
    vc.fill = hdr_fill("F8FAFC")
    vc.alignment = center()
    vc.border = thin_border()

    # Change row
    ws1.merge_cells(start_row=base_row+3, start_column=col_start, end_row=base_row+3, end_column=col_start+2)
    cc = ws1.cell(row=base_row+3, column=col_start, value=change)
    cc.font = Font(name="Arial", italic=True, size=8,
                   color=C["accent"] if '+' in str(change) else C["gray"])
    cc.fill = hdr_fill("F0F4F8")
    cc.alignment = center()

for r in [6,7,8,9,10,11,12,13,14,15]: ws1.row_dimensions[r].height = 22

# Regional performance table
row = 19
ws1.merge_cells(f"A{row}:J{row}")
ws1[f"A{row}"].value = "REGIONAL PERFORMANCE BREAKDOWN — FY2024"
ws1[f"A{row}"].font  = bold_white(10)
ws1[f"A{row}"].fill  = hdr_fill(C["header"])
ws1[f"A{row}"].alignment = center()
ws1.row_dimensions[row].height = 24

row += 1
headers = [("Region",14),("Revenue ($)",16),("Gross Profit ($)",16),("GP Margin %",12),
           ("Transactions",13),("Unique Customers",15),("Avg Order Value ($)",16),
           ("YoY Rev Growth",13),("Top Category",18),("Stores",8)]
write_header_row(ws1, row, headers, C["accent2"])

region_data = fy24.merge(stores[['store_id','region']], on='store_id')
region_data = region_data.merge(products[['product_id','category']], on='product_id')
rg = region_data.groupby('region').agg(
    revenue=('total_revenue','sum'), profit=('gross_profit','sum'),
    txns=('transaction_id','count'), custs=('customer_id','nunique'),
    aov=('total_revenue','mean'),
).reset_index()
rg['margin'] = rg['profit'] / rg['revenue']
top_cat = region_data.groupby(['region','category'])['total_revenue'].sum().reset_index()
top_cat = top_cat.loc[top_cat.groupby('region')['total_revenue'].idxmax()][['region','category']]
rg = rg.merge(top_cat, on='region')

fy23_r = trans[trans['date'].dt.year==2023].merge(stores[['store_id','region']], on='store_id')
fy23_rg = fy23_r.groupby('region')['total_revenue'].sum().reset_index().rename(columns={'total_revenue':'rev23'})
rg = rg.merge(fy23_rg, on='region')
rg['yoy'] = (rg['revenue'] - rg['rev23']) / rg['rev23']

store_cnt = stores.groupby('region')['store_id'].count().reset_index()
rg = rg.merge(store_cnt, on='region')

for i, (_, r) in enumerate(rg.iterrows()):
    row += 1
    fmts = ['@','$#,##0','$#,##0','0.0%','#,##0','#,##0','$#,##0.00','+0.0%;-0.0%','@','#,##0']
    write_data_row(ws1, row,
        [r['region'], r['revenue'], r['profit'], r['margin'],
         r['txns'], r['custs'], r['aov'], r['yoy'], r['category'], r['store_id']],
        even=(i%2==0), formats=fmts)

# Totals
row += 1
totals = [("TOTAL", rg['revenue'].sum(), rg['profit'].sum(),
           rg['profit'].sum()/rg['revenue'].sum(), rg['txns'].sum(),
           rg['custs'].sum(), rg['aov'].mean(), None, "", rg['store_id'].sum())]
for (label, *vals) in totals:
    ws1.cell(row=row, column=1, value=label).font = bold_dark(10)
    fmts2 = ['$#,##0','$#,##0','0.0%','#,##0','#,##0','$#,##0.00','','','#,##0']
    for col, (val, fmt) in enumerate(zip(vals, fmts2), 2):
        c = ws1.cell(row=row, column=col, value=val)
        c.font = bold_dark(10); c.fill = hdr_fill("E8F5E9")
        c.number_format = fmt; c.border = thin_border()
        c.alignment = right()

for col in range(1, 11):
    set_col_width(ws1, col, [14,16,16,12,13,15,16,13,18,8][col-1])

# ══════════════════════════════════════════════════════════════
# SHEET 2: MONTHLY KPI DASHBOARD
# ══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("📅 Monthly KPIs")
ws2.sheet_view.showGridLines = False

ws2.merge_cells("A1:K1")
ws2["A1"].value = "NOVAMART — MONTHLY KPI TRACKER  |  2022–2024"
ws2["A1"].font  = bold_white(14)
ws2["A1"].fill  = hdr_fill(C["dark_bg"])
ws2["A1"].alignment = center()
ws2.row_dimensions[1].height = 40

monthly = trans.copy()
monthly['year'] = monthly['date'].dt.year
monthly['month']= monthly['date'].dt.month
monthly['month_name'] = monthly['date'].dt.strftime('%b')

mg = monthly.groupby(['year','month','month_name']).agg(
    revenue=('total_revenue','sum'), profit=('gross_profit','sum'),
    txns=('transaction_id','count'), custs=('customer_id','nunique'),
    aov=('total_revenue','mean')
).reset_index().sort_values(['year','month'])
mg['margin'] = mg['profit'] / mg['revenue']
mg['label']  = mg['month_name'] + ' ' + mg['year'].astype(str)

row = 3
headers = [
    ("Period",12),("Year",7),("Month",8),("Revenue ($)",15),("Gross Profit ($)",15),
    ("GP Margin %",12),("Transactions",13),("Unique Customers",15),
    ("Avg Order Value ($)",16),("MoM Growth %",12),("QoQ Revenue",13)
]
write_header_row(ws2, row, headers, C["header"])

prev_rev = None
for i, (_, r) in enumerate(mg.iterrows()):
    row += 1
    mom = (r['revenue']-prev_rev)/prev_rev if prev_rev else None
    fmts = ['@','#,##0','#,##0','$#,##0','$#,##0','0.0%','#,##0','#,##0','$#,##0.00','+0.0%;[Red]-0.0%','@']
    write_data_row(ws2, row,
        [r['label'], r['year'], r['month'], r['revenue'], r['profit'],
         r['margin'], r['txns'], r['custs'], r['aov'], mom, ""],
        even=(i%2==0), formats=fmts)
    prev_rev = r['revenue']

# Colour scale on Revenue column
from openpyxl.formatting.rule import ColorScaleRule
cs_rule = ColorScaleRule(start_type='min', start_color='FEF2F2',
                          mid_type='percentile', mid_value=50, mid_color='FFFBEB',
                          end_type='max', end_color='ECFDF5')
ws2.conditional_formatting.add(f"D4:D{row}", cs_rule)

# Add a line chart for revenue
chart = LineChart()
chart.title = "Monthly Revenue Trend"
chart.style = 10
chart.y_axis.title = "Revenue ($)"
chart.x_axis.title = "Period"
chart.height = 12; chart.width = 22

data_ref = Reference(ws2, min_col=4, min_row=3, max_row=row)
cat_ref  = Reference(ws2, min_col=1, min_row=4, max_row=row)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cat_ref)
chart.series[0].graphicalProperties.line.solidFill = "00C97A"
chart.series[0].graphicalProperties.line.width = 20000
ws2.add_chart(chart, f"A{row+3}")

for col, width in enumerate([12,7,8,15,15,12,13,15,16,12,13], 1):
    set_col_width(ws2, col, width)

# ══════════════════════════════════════════════════════════════
# SHEET 3: PRODUCT PERFORMANCE
# ══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("📦 Product Performance")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:I1")
ws3["A1"].value = "PRODUCT PERFORMANCE ANALYSIS — ALL YEARS"
ws3["A1"].font  = bold_white(13)
ws3["A1"].fill  = hdr_fill(C["dark_bg"])
ws3["A1"].alignment = center()
ws3.row_dimensions[1].height = 36

prod_perf = trans.merge(products[['product_id','product_name','category','subcategory','brand']], on='product_id')
pp = prod_perf.groupby(['product_name','category','subcategory','brand']).agg(
    units=('quantity','sum'), revenue=('total_revenue','sum'),
    profit=('gross_profit','sum'), txns=('transaction_id','count')
).reset_index()
pp['margin']  = pp['profit'] / pp['revenue']
pp['avg_price'] = pp['revenue'] / pp['units']
pp['rank']    = pp['revenue'].rank(ascending=False, method='min').astype(int)
pp = pp.sort_values('revenue', ascending=False).head(50)

row = 3
headers = [
    ("Rank",6),("Product Name",32),("Category",18),("Subcategory",16),("Brand",14),
    ("Units Sold",12),("Revenue ($)",14),("Gross Profit ($)",14),("GP Margin %",12)
]
write_header_row(ws3, row, headers, C["accent"] + "CC")

for i, (_, r) in enumerate(pp.iterrows()):
    row += 1
    fmts = ['#,##0','@','@','@','@','#,##0','$#,##0','$#,##0','0.0%']
    write_data_row(ws3, row,
        [r['rank'], r['product_name'], r['category'], r['subcategory'], r['brand'],
         r['units'], r['revenue'], r['profit'], r['margin']],
        even=(i%2==0), formats=fmts)

# Pie chart by category
cat_rev = trans.merge(products[['product_id','category']], on='product_id')
cr = cat_rev.groupby('category')['total_revenue'].sum().reset_index()
cr = cr.sort_values('total_revenue', ascending=False)

ws3.cell(row=row+3, column=1).value = "REVENUE BY CATEGORY"
ws3.cell(row=row+3, column=1).font  = bold_dark(11)

cat_start = row + 4
for i, (_, r) in enumerate(cr.iterrows()):
    ws3.cell(row=cat_start+i, column=1).value = r['category']
    ws3.cell(row=cat_start+i, column=2).value = r['total_revenue']
    ws3.cell(row=cat_start+i, column=2).number_format = '$#,##0'

pie = PieChart()
pie.title = "Revenue by Category"
pie.style = 10; pie.height = 11; pie.width = 14
data_ref2 = Reference(ws3, min_col=2, min_row=cat_start, max_row=cat_start+len(cr)-1)
cat_ref2  = Reference(ws3, min_col=1, min_row=cat_start, max_row=cat_start+len(cr)-1)
pie.add_data(data_ref2)
pie.set_categories(cat_ref2)
ws3.add_chart(pie, f"D{row+3}")

for col, w in enumerate([6,32,18,16,14,12,14,14,12], 1):
    set_col_width(ws3, col, w)

# ══════════════════════════════════════════════════════════════
# SHEET 4: CUSTOMER INTELLIGENCE
# ══════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("👥 Customer Intelligence")
ws4.sheet_view.showGridLines = False

ws4.merge_cells("A1:H1")
ws4["A1"].value = "CUSTOMER INTELLIGENCE — RFM SEGMENTATION & CHURN RISK"
ws4["A1"].font  = bold_white(12)
ws4["A1"].fill  = hdr_fill(C["dark_bg"])
ws4["A1"].alignment = center()
ws4.row_dimensions[1].height = 36

# RFM + churn data
rfm_full = rfm.merge(churn, on='customer_id', how='left')
rfm_full = rfm_full.merge(customers[['customer_id','first_name','last_name','loyalty_tier','age']], on='customer_id', how='left')
rfm_full['customer_name'] = rfm_full['first_name'] + ' ' + rfm_full['last_name']
rfm_full['churn_probability'] = rfm_full['churn_probability'].fillna(0)

# Segment summary
seg_sum = rfm_full.groupby('segment_x').agg(
    count=('customer_id','count'), avg_monetary=('monetary','mean'),
    avg_recency=('recency','mean'), avg_frequency=('frequency','mean'),
    high_risk=('churn_risk_label', lambda x: (x=='High Risk').sum())
).reset_index().rename(columns={'segment_x':'segment'})

row = 3
ws4.merge_cells(f"A{row}:H{row}")
ws4[f"A{row}"].value = "SEGMENT SUMMARY"
ws4[f"A{row}"].font  = bold_white(10)
ws4[f"A{row}"].fill  = hdr_fill(C["accent2"])
ws4[f"A{row}"].alignment = center()

row += 1
headers = [("Segment",18),("Customers",12),("Avg Revenue ($)",15),("Avg Recency (days)",18),
           ("Avg Frequency",14),("High Churn Risk",15),("Action",28),("Priority",10)]
write_header_row(ws4, row, headers, C["header"])

actions = {
    'Champions':         ('Reward & retain — loyalty perks', 'HIGH'),
    'Loyal Customers':   ('Upsell premium products',          'MED'),
    'Potential Loyalists':('Convert with targeted offers',    'MED'),
    'At Risk':           ('Win-back email campaign NOW',       'URGENT'),
    'Hibernating':       ('Re-engagement offer or sunset',    'LOW'),
}
for i, (_, r) in enumerate(seg_sum.sort_values('avg_monetary', ascending=False).iterrows()):
    row += 1
    action, priority = actions.get(r['segment'], ('Review', 'LOW'))
    bg = C["red_bg"] if priority == 'URGENT' else (C["green_bg"] if priority == 'HIGH' else C["white"])
    fmts = ['@','#,##0','$#,##0.00','#,##0.0','#,##0.0','#,##0','@','@']
    write_data_row(ws4, row,
        [r['segment'], r['count'], r['avg_monetary'], r['avg_recency'],
         r['avg_frequency'], r['high_risk'], action, priority],
        even=(i%2==0), formats=fmts)
    for col in range(1,9):
        ws4.cell(row=row, column=col).fill = hdr_fill(bg)

# Top at-risk high-value customers
row += 2
ws4.merge_cells(f"A{row}:H{row}")
ws4[f"A{row}"].value = "HIGH-VALUE AT-RISK CUSTOMERS (Top 20)"
ws4[f"A{row}"].font  = bold_white(10)
ws4[f"A{row}"].fill  = hdr_fill(C["danger"])
ws4[f"A{row}"].alignment = center()

row += 1
headers2 = [("Customer Name",22),("Loyalty Tier",12),("Age",8),("Last Purchase (days)",18),
            ("Lifetime Revenue ($)",18),("Frequency",11),("Churn Prob %",13),("Risk Level",12)]
write_header_row(ws4, row, headers2, C["header"])

at_risk = rfm_full[rfm_full['churn_risk_label']=='High Risk'].sort_values('monetary', ascending=False).head(20)
for i, (_, r) in enumerate(at_risk.iterrows()):
    row += 1
    fmts = ['@','@','#,##0','#,##0','$#,##0','#,##0','0.0%','@']
    write_data_row(ws4, row,
        [r.get('customer_name','N/A'), r.get('loyalty_tier','N/A'), r.get('age',0),
         r['recency'], r['monetary'], r['frequency'],
         r['churn_probability'], r.get('churn_risk_label','N/A')],
        even=(i%2==0), formats=fmts)
    # Red highlight for churn prob
    cp_cell = ws4.cell(row=row, column=7)
    if r['churn_probability'] > 0.7:
        cp_cell.fill = hdr_fill("FECACA")
        cp_cell.font = Font(name="Arial", color=C["danger"], bold=True, size=9)

for col, w in enumerate([22,12,8,18,18,11,13,12], 1):
    set_col_width(ws4, col, w)

# ══════════════════════════════════════════════════════════════
# SHEET 5: INVENTORY STATUS
# ══════════════════════════════════════════════════════════════
inv  = pd.read_csv(f"{DATA}/inventory.csv")
ws5  = wb.create_sheet("📋 Inventory Status")
ws5.sheet_view.showGridLines = False

ws5.merge_cells("A1:G1")
ws5["A1"].value = "INVENTORY STATUS REPORT — CURRENT SNAPSHOT"
ws5["A1"].font  = bold_white(12)
ws5["A1"].fill  = hdr_fill(C["dark_bg"])
ws5["A1"].alignment = center()
ws5.row_dimensions[1].height = 36

# Summary stats
total_skus    = len(inv)
low_stock     = (inv['status']=='Low Stock').sum()
out_of_stock  = (inv['status']=='Out of Stock').sum()
in_stock      = (inv['status']=='In Stock').sum()

row = 3
for i, (label, val, color) in enumerate([
    ("Total SKU-Store Combos", total_skus, C["accent2"]),
    ("In Stock", in_stock, C["accent"]),
    ("Low Stock ⚠️", low_stock, C["warning"]),
    ("Out of Stock 🔴", out_of_stock, C["danger"]),
], 0):
    col_s = i*2 + 1
    ws5.merge_cells(start_row=3, start_column=col_s, end_row=3, end_column=col_s+1)
    ws5.merge_cells(start_row=4, start_column=col_s, end_row=4, end_column=col_s+1)
    lc = ws5.cell(row=3, column=col_s, value=label)
    lc.font = bold_white(9); lc.fill = hdr_fill(color); lc.alignment = center()
    vc = ws5.cell(row=4, column=col_s, value=val)
    vc.font = Font(name="Arial", bold=True, size=16, color=color)
    vc.fill = hdr_fill("F8FAFC"); vc.alignment = center()
    vc.number_format = '#,##0'; vc.border = thin_border()

inv_detail = inv.merge(stores[['store_id','store_name','region']], on='store_id')
inv_detail = inv_detail.merge(products[['product_id','product_name','category']], on='product_id')
inv_alert  = inv_detail[inv_detail['status'].isin(['Low Stock','Out of Stock'])].sort_values('qty_on_hand')

row = 6
ws5.merge_cells(f"A{row}:G{row}")
ws5[f"A{row}"].value = f"REORDER ALERTS — {len(inv_alert)} Items Requiring Attention"
ws5[f"A{row}"].font  = bold_white(10)
ws5[f"A{row}"].fill  = hdr_fill(C["danger"])
ws5[f"A{row}"].alignment = center()

row += 1
headers = [("Store",22),("Region",10),("Product",30),("Category",18),
           ("Qty On Hand",13),("Reorder Level",13),("Status",14)]
write_header_row(ws5, row, headers, C["header"])

for i, (_, r) in enumerate(inv_alert.head(60).iterrows()):
    row += 1
    fmts = ['@','@','@','@','#,##0','#,##0','@']
    write_data_row(ws5, row,
        [r['store_name'], r['region'], r['product_name'], r['category'],
         r['qty_on_hand'], r['reorder_level'], r['status']],
        even=(i%2==0), formats=fmts)
    status_cell = ws5.cell(row=row, column=7)
    if r['status'] == 'Out of Stock':
        status_cell.fill = hdr_fill("FEE2E2")
        status_cell.font = Font(name="Arial", color=C["danger"], bold=True, size=9)
    else:
        status_cell.fill = hdr_fill("FEF3C7")
        status_cell.font = Font(name="Arial", color="B45309", bold=True, size=9)

for col, w in enumerate([22,10,30,18,13,13,14], 1):
    set_col_width(ws5, col, w)

# ══════════════════════════════════════════════════════════════
# SHEET 6: SQL QUERY LIBRARY
# ══════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("🗄️ SQL Query Library")
ws6.sheet_view.showGridLines = False

ws6.merge_cells("A1:C1")
ws6["A1"].value = "SQL QUERY REFERENCE LIBRARY"
ws6["A1"].font  = bold_white(13)
ws6["A1"].fill  = hdr_fill(C["dark_bg"])
ws6["A1"].alignment = center()
ws6.row_dimensions[1].height = 36

queries = [
    ("Monthly Revenue by Region",
     "SELECT d.year, d.month, s.region, SUM(t.total_revenue) AS revenue\nFROM fact_transactions t\nJOIN dim_date d ON t.date = d.date_id\nJOIN dim_stores s ON t.store_id = s.store_id\nGROUP BY d.year, d.month, s.region\nORDER BY d.year, d.month, s.region;",
     "Power BI: Monthly Revenue Trend page"),
    ("Top Products with Ranking",
     "SELECT RANK() OVER (ORDER BY SUM(total_revenue) DESC) AS rank,\n  p.product_name, p.category, SUM(total_revenue) AS revenue\nFROM fact_transactions t\nJOIN dim_products p ON t.product_id = p.product_id\nGROUP BY p.product_name, p.category\nORDER BY revenue DESC LIMIT 20;",
     "Product Performance sheet"),
    ("Customer RFM Scores",
     "SELECT customer_id,\n  CURRENT_DATE - MAX(date) AS recency,\n  COUNT(DISTINCT transaction_id) AS frequency,\n  SUM(total_revenue) AS monetary\nFROM fact_transactions\nGROUP BY customer_id;",
     "Customer Intelligence sheet"),
    ("Inventory Reorder Alert",
     "SELECT s.store_name, p.product_name, i.qty_on_hand, i.reorder_level\nFROM fact_inventory i\nJOIN dim_stores s ON i.store_id = s.store_id\nJOIN dim_products p ON i.product_id = p.product_id\nWHERE i.qty_on_hand <= i.reorder_level\nORDER BY qty_on_hand ASC;",
     "Inventory Status sheet"),
    ("YoY Revenue Growth",
     "SELECT year, month,\n  SUM(revenue) AS rev,\n  LAG(SUM(revenue)) OVER (PARTITION BY month ORDER BY year) AS prev_rev,\n  ROUND((SUM(revenue)-LAG(SUM(revenue)) OVER (PARTITION BY month ORDER BY year))\n    / LAG(SUM(revenue)) OVER (PARTITION BY month ORDER BY year)*100,2) AS yoy_pct\nFROM vw_monthly_revenue\nGROUP BY year, month ORDER BY year, month;",
     "Executive Summary KPIs"),
]

row = 3
write_header_row(ws6, row, [("Query Name",25),("SQL Code",70),("Used In",30)], C["accent2"])
for i, (name, sql, usage) in enumerate(queries):
    row += 1
    ws6.row_dimensions[row].height = max(60, sql.count('\n') * 15)
    nc = ws6.cell(row=row, column=1, value=name)
    nc.font = Font(name="Arial", bold=True, size=9, color=C["header"]); nc.border = thin_border()
    nc.fill = hdr_fill(C["blue_bg"] if i%2==0 else C["white"])
    nc.alignment = Alignment(vertical='top', wrap_text=True)

    sc = ws6.cell(row=row, column=2, value=sql)
    sc.font = Font(name="Courier New", size=8, color="1A1A3E")
    sc.fill = hdr_fill("F0F4FF"); sc.border = thin_border()
    sc.alignment = Alignment(vertical='top', wrap_text=True)

    uc = ws6.cell(row=row, column=3, value=usage)
    uc.font = normal(9, C["gray"]); uc.border = thin_border()
    uc.fill = hdr_fill(C["white"]); uc.alignment = Alignment(vertical='top', wrap_text=True)

set_col_width(ws6, 1, 25)
set_col_width(ws6, 2, 70)
set_col_width(ws6, 3, 30)

# ─── SAVE ───────────────────────────────────────────────────
out_path = f"{OUT}/RetailPulse360_Management_Report.xlsx"
wb.save(out_path)
print(f"✅ Excel report saved: {out_path}")
print(f"   Sheets: {[ws.title for ws in wb.worksheets]}")
